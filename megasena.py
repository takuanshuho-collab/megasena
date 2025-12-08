"""
Script para coletar os últimos 100 resultados da Mega-Sena na API oficial da CAIXA,
calcular a frequência das dezenas (1 a 60), gerar (opcionalmente) gráfico de frequência e
criar jogos de 6 números com regras parametrizáveis via CLI:
- Anti-sequência (comprimento máximo de sequência consecutiva)
- Paridade (alvo de pares e tolerância)
- Mesmo final (máximo de dezenas com o mesmo dígito final)
- Soma total (intervalo permitido)
- Intervalo mínimo (min_span) entre a menor e a maior dezena (evita jogos muito concentrados)
- Máximo de dezenas no mesmo grupo de dezenas (max_same_tens, ex.: 10–19)

Funcionalidades adicionais:
- Cache local em JSON para reduzir chamadas à API / reaproveitar dados
- Exportação: CSV (frequências e jogos) e Excel (.xlsx) com abas Frequencias, Jogos, Parametros, Historico, Metadados
- Gráfico embutido na aba Frequencias (openpyxl)
- Flag --no-chart para desabilitar gráfico (PNG e também o gráfico embutido)
- Flag --export-dir para salvar freq.csv, jogos.csv, gráfico e Excel em subpasta com timestamp

Dependências: requests, beautifulsoup4, pytest, matplotlib (opcional), pandas (engine openpyxl)
"""
from __future__ import annotations
import logging
import random
import re
import argparse
import json
import os
import csv
from dataclasses import dataclass
from typing import Dict, List, Set, Tuple
from datetime import datetime
from pathlib import Path

import requests
from bs4 import BeautifulSoup

API_BASE = "https://servicebus2.caixa.gov.br/portaldeloterias/api/megasena"
TIMEOUT = 10
CACHE_FILE = "cache_megasena.json"
FETCH_SOURCE = "unknown"
LAST_CONCURSO = None

logging.basicConfig(level=logging.INFO, format="[%(levelname)s] %(message)s")

@dataclass
class FetchResult:
    draws: List[List[int]]
    history_set: Set[Tuple[int, ...]]


def _safe_get_json(url: str) -> dict | None:
    try:
        logging.debug(f"GET {url}")
        resp = requests.get(url, timeout=TIMEOUT)
        if resp.status_code != 200:
            logging.warning(f"Falha HTTP {resp.status_code} ao acessar {url}")
            return None
        return resp.json()
    except requests.exceptions.RequestException as e:
        logging.error(f"Erro de conexão ao acessar {url}: {e}")
        return None
    except ValueError as e:
        logging.error(f"Erro ao interpretar JSON de {url}: {e}")
        return None


def _save_cache(draws: List[List[int]], meta: dict | None = None) -> None:
    try:
        payload = {
            "fetched_at": datetime.now().isoformat(),
            "draws": draws,
            "meta": meta or {},
        }
        with open(CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False)
        logging.info(f"Cache salvo em {CACHE_FILE}")
    except Exception as e:
        logging.warning(f"Falha ao salvar cache: {e}")


def _load_cache() -> FetchResult | None:
    if not os.path.exists(CACHE_FILE):
        return None
    try:
        with open(CACHE_FILE, "r", encoding="utf-8") as f:
            payload = json.load(f)
        draws = payload.get("draws") or []
        meta = payload.get("meta") or {}
        # Atualiza ultimo concurso a partir do cache, se disponível
        try:
            ln = meta.get("last_number")
            if ln is not None:
                global LAST_CONCURSO
                LAST_CONCURSO = int(ln)
        except Exception:
            pass
        history_set = {tuple(sorted(x)) for x in draws if isinstance(x, list) and len(x) == 6}
        if draws:
            logging.info(f"Usando cache local ({len(draws)} concursos)")
            return FetchResult(draws=draws, history_set=history_set)
    except Exception as e:
        logging.warning(f"Falha ao ler cache: {e}")
    return None


def fetch_data() -> FetchResult:
    global FETCH_SOURCE
    logging.info("Iniciando busca dos últimos resultados (API CAIXA)")
    latest = _safe_get_json(API_BASE)
    draws: List[List[int]] = []
    history_set: Set[Tuple[int, ...]] = set()

    if latest and isinstance(latest, dict) and "numero" in latest:
        try:
            last_number = int(latest["numero"])  # número do último concurso
        except (ValueError, TypeError):
            last_number = None
        if last_number:
            logging.info(f"Último concurso na API: {last_number}")
            for n in range(last_number, max(last_number - 100, 0), -1):
                url = f"{API_BASE}/{n}"
                data_n = _safe_get_json(url)
                if not data_n:
                    continue
                dezenas = data_n.get("listaDezenas") or []
                try:
                    nums = [int(d) for d in dezenas]
                except Exception:
                    nums = []
                if len(nums) == 6 and all(1 <= x <= 60 for x in nums):
                    draws.append(nums)
                    history_set.add(tuple(sorted(nums)))
            if draws:
                logging.info(f"Coleta via API concluída: {len(draws)} concursos")
                _save_cache(draws, meta={"last_number": last_number})
                FETCH_SOURCE = "api"
                global LAST_CONCURSO
                LAST_CONCURSO = last_number
                return FetchResult(draws=draws, history_set=history_set)
        else:
            logging.warning("Campo 'numero' ausente ou inválido no JSON da API base.")
    else:
        logging.warning("Falhou obter o último concurso na API base.")

    # Tenta cache
    cache = _load_cache()
    if cache:
        FETCH_SOURCE = "cache"
        return cache

    # Fallback: scraping
    logging.info("API indisponível e sem cache válido. Iniciando fallback por scraping (BeautifulSoup)")
    FETCH_SOURCE = "scraping"
    return _fallback_scraping()


def _fallback_scraping() -> FetchResult:
    draws: List[List[int]] = []
    history_set: Set[Tuple[int, ...]] = set()
    urls = [
        "https://www.megasena.com/resultados",
        "https://www.mazusoft.com.br/mega/resultados.php",
    ]
    pattern = re.compile(r"\b([0-5]?\d|60)\b")

    for url in urls:
        try:
            logging.info(f"Scraping: {url}")
            r = requests.get(url, timeout=TIMEOUT)
            if r.status_code != 200:
                logging.warning(f"HTTP {r.status_code} ao scrapear {url}")
                continue
            soup = BeautifulSoup(r.text, "html.parser")
            texts = soup.stripped_strings
            buffer: List[int] = []
            for token in texts:
                for match in pattern.findall(token):
                    try:
                        num = int(match)
                    except Exception:
                        continue
                    if 1 <= num <= 60:
                        buffer.append(num)
                if len(buffer) >= 6:
                    nums = buffer[-6:]
                    if len(nums) == 6:
                        tup = tuple(sorted(nums))
                        if tup not in history_set:
                            draws.append(nums)
                            history_set.add(tup)
                    if len(draws) >= 100:
                        break
            if len(draws) >= 100:
                break
        except requests.exceptions.RequestException as e:
            logging.error(f"Erro de conexão no scraping {url}: {e}")
            continue
    logging.info(f"Scraping concluído com {len(draws)} concursos")
    if draws:
        _save_cache(draws, meta={"source": "scraping"})
    return FetchResult(draws=draws, history_set=history_set)


def calculate_frequency(draws: List[List[int]]) -> Dict[int, int]:
    logging.info("Processando frequências das dezenas")
    freq = {i: 0 for i in range(1, 61)}
    for concurso in draws:
        for d in concurso:
            if 1 <= d <= 60:
                freq[d] += 1
    return freq


def plot_frequency(freq: Dict[int, int], filename: str = "freq_megasena.png") -> str:
    import matplotlib.pyplot as plt
    xs = list(range(1, 61))
    ys = [freq[i] for i in xs]
    plt.figure(figsize=(14, 6))
    plt.bar(xs, ys, color="steelblue")
    plt.title("Frequência das dezenas (últimos concursos)")
    plt.xlabel("Dezena")
    plt.ylabel("Frequência")
    plt.xticks(xs, [f"{i:02d}" for i in xs], rotation=0, fontsize=8)
    plt.grid(axis="y", linestyle=":", alpha=0.4)
    plt.tight_layout()
    plt.savefig(filename, dpi=200)
    plt.close()
    logging.info(f"Gráfico salvo em: {filename}")
    return filename


def export_frequency_csv(freq: Dict[int, int], filename: str) -> str:
    rows = [(i, freq[i]) for i in range(1, 61)]
    try:
        with open(filename, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["dezena", "frequencia"])
            for d, c in rows:
                writer.writerow([d, c])
        logging.info(f"CSV de frequência salvo em: {filename}")
        return filename
    except Exception as e:
        logging.error(f"Falha ao salvar CSV de frequência: {e}")
        return filename


def export_games_csv(
    games: List[List[int]],
    filename: str,
    params: dict,
) -> str:
    ts = datetime.now().isoformat()
    try:
        with open(filename, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            header = [
                "timestamp", "game_index", "numbers",
                "mix", "max_run", "target_even", "tolerance",
                "max_same_ending", "min_sum", "max_sum", "min_span", "max_same_tens",
                "seed", "n_games"
            ]
            writer.writerow(header)
            for idx, g in enumerate(games, start=1):
                numbers_str = " ".join(f"{n:02d}" for n in g)
                writer.writerow([
                    ts, idx, numbers_str,
                    params.get("mix"), params.get("max_run"), params.get("target_even"), params.get("tolerance"),
                    params.get("max_same_ending"), params.get("min_sum"), params.get("max_sum"), params.get("min_span"), params.get("max_same_tens"),
                    params.get("seed"), params.get("n_games"),
                ])
        logging.info(f"CSV de jogos salvo em: {filename}")
        return filename
    except Exception as e:
        logging.error(f"Falha ao salvar CSV de jogos: {e}")
        return filename


def export_excel(freq: Dict[int, int], games: List[List[int]], filename: str, params: dict, draws: List[List[int]] | None = None, metadata: Dict[str, str] | None = None, add_chart: bool = True) -> str:
    """Exporta um Excel (.xlsx) com abas: Frequencias, Jogos, Parametros, Historico e Metadados.
    Adiciona gráfico de barras na aba Frequencias usando openpyxl.
    """
    import pandas as pd
    from openpyxl.chart import BarChart, Reference
    from openpyxl.formatting.rule import ColorScaleRule, Top10Rule
    from openpyxl.styles import PatternFill

    # Frequencias
    df_freq = pd.DataFrame({"dezena": list(range(1,61)), "frequencia": [freq[i] for i in range(1,61)]})
    # Jogos
    ts = datetime.now().isoformat()
    df_games = pd.DataFrame([
        {"timestamp": ts, "game_index": idx, "numbers": " ".join(f"{n:02d}" for n in g)}
        for idx, g in enumerate(games, start=1)
    ])
    # Parametros
    df_params = pd.DataFrame({"parametro": list(params.keys()), "valor": list(params.values())})

    # Historico
    if draws and len(draws) > 0:
        rows = []
        for idx, d in enumerate(draws, start=1):
            row = {"concurso_idx": idx}
            for j, val in enumerate(sorted(d), start=1):
                row[f"D{j}"] = val
            rows.append(row)
        df_hist = pd.DataFrame(rows)
    else:
        df_hist = pd.DataFrame({"info": ["Historico não disponível nesta execução"]})

    # Metadados
    meta = metadata or {}
    base_meta = {
        "coleta_timestamp": ts,
        "total_concursos": len(draws or []),
        "fonte": meta.get("fonte", "desconhecida"),
        "ultimo_concurso": meta.get("ultimo_concurso", ""),
        "observacao": meta.get("observacao", "Resultados coletados dos últimos até 100 concursos"),
    }
    df_meta = pd.DataFrame({"campo": list(base_meta.keys()), "valor": list(base_meta.values())})

    try:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df_freq.to_excel(writer, index=False, sheet_name='Frequencias')
            df_games.to_excel(writer, index=False, sheet_name='Jogos')
            df_params.to_excel(writer, index=False, sheet_name='Parametros')
            df_hist.to_excel(writer, index=False, sheet_name='Historico')
            df_meta.to_excel(writer, index=False, sheet_name='Metadados')

            if add_chart:
                wb = writer.book
                ws = wb['Frequencias']
                data_ref = Reference(ws, min_col=2, min_row=2, max_col=2, max_row=61)
                cats_ref = Reference(ws, min_col=1, min_row=2, max_row=61)
                chart = BarChart()
                chart.title = 'Frequência das dezenas'
                chart.y_axis.title = 'Frequência'
                chart.x_axis.title = 'Dezena'
                chart.add_data(data_ref, titles_from_data=False)
                chart.set_categories(cats_ref)
                ws.add_chart(chart, 'E2')

                # Formatação condicional na coluna de frequências (B2:B61)
                freq_range = 'B2:B61'
                # Escala de cores (min->max)
                color_scale = ColorScaleRule(start_type='min', start_color='FFC7CE',
                                             mid_type='percentile', mid_value=50, mid_color='FFEB84',
                                             end_type='max', end_color='C6EFCE')
                ws.conditional_formatting.add(freq_range, color_scale)
                # Destaque Top 20 (verde)
                top_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                ws.conditional_formatting.add(freq_range, Top10Rule(rank=20, percent=False, bottom=False, fill=top_fill))
                # Destaque Bottom 20 (vermelho)
                bottom_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                ws.conditional_formatting.add(freq_range, Top10Rule(rank=20, percent=False, bottom=True, fill=bottom_fill))

        logging.info(f"Excel salvo em: {filename}")
        return filename
    except Exception as e:
        logging.error(f"Falha ao salvar Excel: {e}")
        return filename


def _split_tiers(freq: Dict[int, int]) -> Tuple[List[int], List[int], List[int]]:
    ordered = sorted(freq.items(), key=lambda kv: (-kv[1], kv[0]))
    numbers_ordered = [n for n, _ in ordered]
    top = numbers_ordered[:20]
    middle = numbers_ordered[20:40]
    bottom = numbers_ordered[40:60]
    return top, middle, bottom


def _max_consecutive_run(nums: List[int]) -> int:
    s = sorted(nums)
    run = 1
    best = 1
    for i in range(1, len(s)):
        if s[i] == s[i-1] + 1:
            run += 1
            best = max(best, run)
        else:
            run = 1
    return best


def _parity_ok(nums: List[int], target_even: int = 3, tolerance: int = 0) -> bool:
    even = sum(1 for x in nums if x % 2 == 0)
    odd = len(nums) - even
    return abs(even - target_even) <= tolerance and abs(odd - (len(nums) - target_even)) <= tolerance


def _same_ending_ok(nums: List[int], max_same_ending: int = 2) -> bool:
    counts = {d: 0 for d in range(10)}
    for x in nums:
        counts[x % 10] += 1
    return max(counts.values()) <= max_same_ending


def _sum_ok(nums: List[int], min_sum: int, max_sum: int) -> bool:
    total = sum(nums)
    return min_sum <= total <= max_sum


def _span_ok(nums: List[int], min_span: int) -> bool:
    s = sorted(nums)
    return (s[-1] - s[0]) >= min_span


def _same_tens_ok(nums: List[int], max_same_tens: int) -> bool:
    buckets = {b: 0 for b in range(6)}  # 0: 0–9, 1:10–19 ... 5:50–59
    for x in nums:
        b = x // 10
        if b > 5:
            b = 5
        buckets[b] += 1
    return max(buckets.values()) <= max_same_tens


def _is_valid_game(
    game: List[int],
    history_set: Set[Tuple[int, ...]],
    max_run: int = 2,
    target_even: int = 3,
    tolerance: int = 0,
    max_same_ending: int = 2,
    min_sum: int = 100,
    max_sum: int = 260,
    min_span: int = 25,
    max_same_tens: int = 3,
) -> bool:
    if len(game) != 6:
        return False
    if len(set(game)) != 6:
        return False
    if not all(1 <= x <= 60 for x in game):
        return False
    if tuple(sorted(game)) in history_set:
        return False
    if _max_consecutive_run(game) > max_run:
        return False
    if not _parity_ok(game, target_even=target_even, tolerance=tolerance):
        return False
    if not _same_ending_ok(game, max_same_ending=max_same_ending):
        return False
    if not _sum_ok(game, min_sum=min_sum, max_sum=max_sum):
        return False
    if not _span_ok(game, min_span=min_span):
        return False
    if not _same_tens_ok(game, max_same_tens=max_same_tens):
        return False
    return True


def generate_games(
    freq: Dict[int, int],
    history_set: Set[Tuple[int, ...]],
    n_games: int = 3,
    seed: int | None = None,
    mix: Tuple[int, int, int] = (3, 2, 1),
    max_run: int = 2,
    target_even: int = 3,
    tolerance: int = 0,
    max_same_ending: int = 2,
    min_sum: int = 100,
    max_sum: int = 260,
    min_span: int = 25,
    max_same_tens: int = 3,
) -> List[List[int]]:
    logging.info("Gerando jogos com regras anti-sequência, paridade, mesmo final, soma, span e grupos de dezenas")
    if seed is not None:
        random.seed(seed)

    top, middle, bottom = _split_tiers(freq)
    games: List[List[int]] = []
    max_attempts = 2000

    for gi in range(n_games):
        attempts = 0
        while attempts < max_attempts:
            attempts += 1
            pick_top = random.sample(top, mix[0])
            pick_mid = random.sample(middle, mix[1])
            pick_bot = random.sample(bottom, mix[2])
            candidate = sorted(pick_top + pick_mid + pick_bot)

            if not _is_valid_game(
                candidate,
                history_set,
                max_run=max_run,
                target_even=target_even,
                tolerance=tolerance,
                max_same_ending=max_same_ending,
                min_sum=min_sum,
                max_sum=max_sum,
                min_span=min_span,
                max_same_tens=max_same_tens,
            ):
                continue
            if tuple(candidate) in {tuple(sorted(g)) for g in games}:
                continue

            games.append(candidate)
            break
        if attempts >= max_attempts and len(games) <= gi:
            logging.error("Falha ao gerar jogo válido após muitas tentativas.")
            raise RuntimeError("Não foi possível gerar jogos válidos dentro do limite de tentativas.")

    return games


def _explain_logic(
    freq: Dict[int, int],
    games: List[List[int]],
    max_run: int,
    target_even: int,
    tolerance: int,
    mix: Tuple[int,int,int],
    max_same_ending: int,
    min_sum: int,
    max_sum: int,
    min_span: int,
    max_same_tens: int,
) -> str:
    top, middle, bottom = _split_tiers(freq)
    msg = [
        "Lógica de geração:",
        "- Frequências calculadas com base nos últimos 100 concursos coletados.",
        f"- Mistura por faixas: TOP/MIDDLE/BOTTOM = {mix}.",
        f"- Anti-sequência: máximo de {max_run} números consecutivos.",
        f"- Paridade: alvo de {target_even} pares (tolerância ±{tolerance}).",
        f"- Mesmo final: no máximo {max_same_ending} dezenas com o mesmo dígito final.",
        f"- Soma total: intervalo permitido {min_sum}–{max_sum}.",
        f"- Intervalo mínimo (span): pelo menos {min_span} entre menor e maior dezena.",
        f"- Grupos de dezenas: no máximo {max_same_tens} números dentro do mesmo grupo (10–19 etc.).",
        "- Jogos não repetem combinações já presentes no histórico carregado e são distintos entre si.",
    ]
    tiers_example = (
        f"Exemplos de faixas: TOP={sorted(top[:6])}, MIDDLE={sorted(middle[:6])}, BOTTOM={sorted(bottom[:6])}"
    )
    msg.append(tiers_example)
    return "\n".join(msg)


def _parse_mix(value: str) -> Tuple[int,int,int]:
    value = value.strip()
    sep = '-' if '-' in value else ','
    parts = value.split(sep)
    if len(parts) != 3:
        raise argparse.ArgumentTypeError("mix deve ter 3 números (ex.: 3-2-1)")
    try:
        t = tuple(int(p) for p in parts)
    except ValueError:
        raise argparse.ArgumentTypeError("mix deve conter inteiros")
    if sum(t) != 6:
        raise argparse.ArgumentTypeError("mix deve somar 6 (ex.: 3-2-1, 2-2-2)")
    if any(x < 0 for x in t):
        raise argparse.ArgumentTypeError("mix não pode ter negativos")
    return t  # type: ignore


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Gerador de jogos da Mega-Sena com regras, gráfico, cache e exportação")
    parser.add_argument("--mix", type=_parse_mix, default=(3,2,1), help="Mistura TOP-MIDDLE-BOTTOM (ex.: 3-2-1 ou 2,2,2)")
    parser.add_argument("--max-run", type=int, default=2, help="Comprimento máximo de sequência consecutiva")
    parser.add_argument("--target-even", type=int, default=3, help="Alvo de dezenas pares (ex.: 3)")
    parser.add_argument("--tolerance", type=int, default=0, help="Tolerância do alvo de pares (ex.: 0)")
    parser.add_argument("--max-same-ending", type=int, default=2, help="Máximo de dezenas com o mesmo dígito final (0..9)")
    parser.add_argument("--min-sum", type=int, default=100, help="Soma mínima permitida do jogo")
    parser.add_argument("--max-sum", type=int, default=260, help="Soma máxima permitida do jogo")
    parser.add_argument("--min-span", type=int, default=25, help="Intervalo mínimo entre menor e maior dezena")
    parser.add_argument("--max-same-tens", type=int, default=3, help="Máximo de dezenas no mesmo grupo (10–19 etc.)")
    parser.add_argument("--seed", type=int, default=None, help="Seed para reprodutibilidade")
    parser.add_argument("--n-games", type=int, default=3, help="Número de jogos a gerar")
    parser.add_argument("--no-chart", action="store_true", help="Não gerar gráfico de frequência")
    parser.add_argument("--freq-csv", type=str, default=None, help="Caminho para salvar CSV de frequência")
    parser.add_argument("--games-csv", type=str, default=None, help="Caminho para salvar CSV dos jogos")
    parser.add_argument("--export-xlsx", type=str, default=None, help="Caminho para salvar Excel (.xlsx) com frequencias, jogos, parametros e historico")
    parser.add_argument("--export-dir", type=str, default=None, help="Diretório para salvar freq.csv, jogos.csv, gráfico e Excel em subpasta com timestamp")
    return parser


def main() -> None:
    parser = build_arg_parser()
    args = parser.parse_args()

    logging.info("===== Mega-Sena (pipeline completo) =====")
    fetch = fetch_data()
    if not fetch.draws:
        logging.error("Nenhum concurso coletado. Encerrando.")
        return

    freq = calculate_frequency(fetch.draws)

    # Prepara diretório de exportação com timestamp
    export_dir = None
    chart_file = None
    if args.export_dir:
        ts_folder = datetime.now().strftime("%Y%m%d_%H%M%S")
        export_dir = Path(args.export_dir) / ts_folder
        export_dir.mkdir(parents=True, exist_ok=True)

    # Gráfico (opcional)
    if not args.no_chart:
        target_chart = str(export_dir / "freq_megasena.png") if export_dir else "freq_megasena.png"
        chart_file = plot_frequency(freq, filename=target_chart)

    # Exporta CSV de frequência
    if export_dir:
        export_frequency_csv(freq, str(export_dir / "freq.csv"))
    elif args.freq_csv:
        export_frequency_csv(freq, args.freq_csv)

    # Gera jogos
    games = generate_games(
        freq,
        fetch.history_set,
        n_games=args.n_games,
        seed=args.seed,
        mix=args.mix,
        max_run=args.max_run,
        target_even=args.target_even,
        tolerance=args.tolerance,
        max_same_ending=args.max_same_ending,
        min_sum=args.min_sum,
        max_sum=args.max_sum,
        min_span=args.min_span,
        max_same_tens=args.max_same_tens,
    )

    print("\n=== Jogos sugeridos (6 dezenas cada) ===")
    for idx, g in enumerate(games, start=1):
        print(f"Jogo {idx}: {' '.join(f'{n:02d}' for n in g)}")

    print("\n=== Explicação da lógica usada ===")
    print(_explain_logic(
        freq,
        games,
        max_run=args.max_run,
        target_even=args.target_even,
        tolerance=args.tolerance,
        mix=args.mix,
        max_same_ending=args.max_same_ending,
        min_sum=args.min_sum,
        max_sum=args.max_sum,
        min_span=args.min_span,
        max_same_tens=args.max_same_tens,
    ))
    if chart_file:
        print(f"\nGráfico de frequência salvo em: {chart_file}")

    # Metadados para Excel
    excel_meta = {
        "fonte": FETCH_SOURCE,
        "ultimo_concurso": str(LAST_CONCURSO or ""),
        "observacao": "Resultados coletados dos últimos até 100 concursos",
    }

    # Exporta CSV / Excel
    params = {
        "mix": args.mix,
        "max_run": args.max_run,
        "target_even": args.target_even,
        "tolerance": args.tolerance,
        "max_same_ending": args.max_same_ending,
        "min_sum": args.min_sum,
        "max_sum": args.max_sum,
        "min_span": args.min_span,
        "max_same_tens": args.max_same_tens,
        "seed": args.seed,
        "n_games": args.n_games,
    }
    if export_dir:
        export_games_csv(games, str(export_dir / "jogos.csv"), params)
        export_excel(freq, games, str(export_dir / "megasena.xlsx"), params, draws=fetch.draws, metadata=excel_meta, add_chart=(not args.no_chart))
    else:
        if args.games_csv:
            export_games_csv(games, args.games_csv, params)
        if args.export_xlsx:
            export_excel(freq, games, args.export_xlsx, params, draws=fetch.draws, metadata=excel_meta, add_chart=(not args.no_chart))


if __name__ == "__main__":
    main()
